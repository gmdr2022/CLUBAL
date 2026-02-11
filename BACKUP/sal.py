# sal.py
# SAL - SESI Agenda Live (v2 UI)
# Layout: Left column (AGORA top, PRÓXIMAS bottom). Right column (Horários rotativos + Clima com ícone).
# Assets: graphics/logo_day(.png), graphics/logo_night(.png), graphics/*.png (icons)
# Logs (gravável): %LOCALAPPDATA%/SAL_SESI_Agenda_Live/logs/sal.log
# Cache clima (gravável): %LOCALAPPDATA%/SAL_SESI_Agenda_Live/package/weather_cache.json (+ package/cache_old/*.json)
# Excel: grade.xlsx (no APP_DIR)

from __future__ import annotations

import os
import sys
import time
import traceback
import threading
from dataclasses import dataclass
from typing import List, Optional, Tuple

import tkinter as tk
from openpyxl import load_workbook

import weather as weather_mod

from datetime import datetime, timedelta, date as dt_date, time as dt_time


# -------------------------
# Paths / Logging
# -------------------------

def app_dir() -> str:
    # Works for python and PyInstaller
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def data_dir(app_name: str = "SAL_SESI_Agenda_Live") -> str:
    """
    Diretório gravável por usuário.
    - Em PCs corporativos, evita problemas quando o app roda em Program Files.
    - Se LOCALAPPDATA não existir, cai no APP_DIR (melhor esforço).
    """
    base = os.environ.get("LOCALAPPDATA")
    if not base:
        p = os.path.join(app_dir(), "_data")
        os.makedirs(p, exist_ok=True)
        return p
    p = os.path.join(base, app_name)
    os.makedirs(p, exist_ok=True)
    return p


APP_DIR = app_dir()
DATA_DIR = data_dir()

GRAPHICS_DIR = os.path.join(APP_DIR, "graphics")

# logs em local gravável
LOGS_DIR = os.path.join(DATA_DIR, "logs")
os.makedirs(LOGS_DIR, exist_ok=True)
LOG_PATH = os.path.join(LOGS_DIR, "sal.log")

# ✅ backend refinamento: rotação/limpeza de logs
LOG_ARCHIVE_DIR = os.path.join(LOGS_DIR, "archive")
os.makedirs(LOG_ARCHIVE_DIR, exist_ok=True)

LOG_ROTATE_MAX_BYTES = 2 * 1024 * 1024   # 2MB
LOG_ARCHIVE_KEEP = 6                     # mantém os últimos N logs arquivados
LOG_ARCHIVE_MAX_AGE_DAYS = 30            # e/ou apaga logs muito antigos


def _safe_unlink(path: str) -> None:
    try:
        os.remove(path)
    except Exception:
        pass


def _rotate_logs_if_needed(logger=None) -> None:
    """
    Rotaciona sal.log quando cresce demais.
    - Move para logs/archive/sal_YYYYMMDD_HHMMSS.log
    - Mantém no máximo LOG_ARCHIVE_KEEP arquivos e remove os muito antigos
    """
    try:
        if not os.path.exists(LOG_PATH):
            return
        size = os.path.getsize(LOG_PATH)
        if size < LOG_ROTATE_MAX_BYTES:
            # também faz limpeza leve do archive
            _cleanup_log_archive()
            return

        ts = time.strftime("%Y%m%d_%H%M%S")
        archived = os.path.join(LOG_ARCHIVE_DIR, f"sal_{ts}.log")
        try:
            os.replace(LOG_PATH, archived)
        except Exception:
            # fallback: copiar truncando (melhor esforço)
            with open(LOG_PATH, "rb") as src, open(archived, "wb") as dst:
                dst.write(src.read())
            with open(LOG_PATH, "w", encoding="utf-8"):
                pass

        if logger:
            logger(f"[LOG] Rotated sal.log -> {archived}")

        _cleanup_log_archive()

    except Exception as e:
        if logger:
            logger(f"[LOG] Rotate error {type(e).__name__}: {e}")


def _cleanup_log_archive() -> None:
    try:
        files = []
        now = time.time()
        max_age = LOG_ARCHIVE_MAX_AGE_DAYS * 86400

        for name in os.listdir(LOG_ARCHIVE_DIR):
            p = os.path.join(LOG_ARCHIVE_DIR, name)
            if not os.path.isfile(p):
                continue
            try:
                st = os.stat(p)
            except Exception:
                continue

            # remove por idade
            if max_age > 0 and (now - st.st_mtime) > max_age:
                _safe_unlink(p)
                continue

            files.append((st.st_mtime, p))

        # remove excedentes por quantidade (mantém os mais recentes)
        files.sort(reverse=True)  # mais novo primeiro
        for _mtime, p in files[LOG_ARCHIVE_KEEP:]:
            _safe_unlink(p)

    except Exception:
        pass


def log(msg: str) -> None:
    try:
        # rotação barata (checa tamanho ocasionalmente)
        _rotate_logs_if_needed()
        ts = time.strftime("%Y-%m-%d %H:%M:%S")
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass


# -------------------------
# Data model
# -------------------------

@dataclass
class ClassItem:
    day: str          # SEG TER QUA QUI SEX SAB DOM
    start: str        # HH:MM
    end: str          # HH:MM
    activity: str
    teacher: str
    location: str
    tag: str          # MENOR/GERAL/...


# -------------------------
# Time parsing helpers (tolerante)
# -------------------------

def parse_hhmm(s: str) -> Optional[int]:
    """
    Aceita:
    - "HH:MM"
    - "HH:MM:SS"
    """
    try:
        s = str(s).strip()
        if not s:
            return None
        parts = s.split(":")
        if len(parts) < 2:
            return None
        hh = int(parts[0])
        mm = int(parts[1])
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return hh * 60 + mm
        return None
    except Exception:
        return None


def _parse_time_obj(s: str) -> Optional[dt_time]:
    m = parse_hhmm(s)
    if m is None:
        return None
    return dt_time(hour=m // 60, minute=m % 60)


def today_3letters_noaccent() -> str:
    wd = time.localtime().tm_wday  # Monday=0
    return ["SEG", "TER", "QUA", "QUI", "SEX", "SAB", "DOM"][wd]


def date_time_strings() -> Tuple[str, str]:
    lt = time.localtime()
    date_s = time.strftime("%d/%m/%Y", lt)
    time_s = time.strftime("%H:%M", lt)
    return date_s, time_s


def theme_is_day() -> bool:
    h = time.localtime().tm_hour
    return 6 <= h <= 17


def fmt_hhmm(mins: Optional[int]) -> str:
    if mins is None:
        return "—"
    hh = mins // 60
    mm = mins % 60
    if mm == 0:
        return f"{hh:02d}H"
    return f"{hh:02d}:{mm:02d}"


# -------------------------
# Excel read
# -------------------------

EXCEL_PATH = os.path.join(APP_DIR, "grade.xlsx")
SHEET_NAME = "SAL"

EXPECTED_HEADERS = ["DIA", "INICIO", "FIM", "ATIVIDADE", "PROFESSOR", "LOCAL", "TAG"]


def _normalize_header(v: object) -> str:
    return str(v or "").strip().upper()


def load_classes_from_excel(path: str) -> List[ClassItem]:
    wb = load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Aba '{SHEET_NAME}' não encontrada.")
    ws = wb[SHEET_NAME]

    # Detecta linha de cabeçalho em até 5 primeiras linhas
    header_row = None
    headers: List[str] = []
    for r in range(1, 6):
        vals = [_normalize_header(ws.cell(row=r, column=c).value) for c in range(1, 30)]
        if "DIA" in vals and "INICIO" in vals and "FIM" in vals:
            header_row = r
            headers = vals
            break
    if header_row is None:
        header_row = 1
        headers = [_normalize_header(ws.cell(row=1, column=c).value) for c in range(1, 30)]

    def col_idx(name: str) -> Optional[int]:
        name = name.upper()
        if name in headers:
            return headers.index(name) + 1
        return None

    c_dia = col_idx("DIA")
    c_ini = col_idx("INICIO")
    c_fim = col_idx("FIM")
    c_ativ = col_idx("ATIVIDADE")
    c_prof = col_idx("PROFESSOR")
    c_loc = col_idx("LOCAL")
    c_tag = col_idx("TAG")

    if not (c_dia and c_ini and c_fim and c_ativ):
        raise RuntimeError("Cabeçalhos necessários não encontrados na aba SAL.")

    items: List[ClassItem] = []
    for r in range(header_row + 1, ws.max_row + 1):
        dia = str(ws.cell(row=r, column=c_dia).value or "").strip().upper()
        ini = str(ws.cell(row=r, column=c_ini).value or "").strip()
        fim = str(ws.cell(row=r, column=c_fim).value or "").strip()
        ativ = str(ws.cell(row=r, column=c_ativ).value or "").strip()

        if not dia or not ini or not fim or not ativ:
            continue

        prof = str(ws.cell(row=r, column=c_prof).value or "").strip() if c_prof else ""
        loc = str(ws.cell(row=r, column=c_loc).value or "").strip() if c_loc else ""
        tag = str(ws.cell(row=r, column=c_tag).value or "").strip().upper() if c_tag else ""

        items.append(ClassItem(dia, ini, fim, ativ, prof, loc, tag))

    return items


# -------------------------
# Day mapping + midnight-safe datetime placement
# -------------------------

DAY_TO_WD = {"SEG": 0, "TER": 1, "QUA": 2, "QUI": 3, "SEX": 4, "SAB": 5, "DOM": 6}


def _best_date_for_daycode(now: datetime, day_code: str) -> Optional[dt_date]:
    """
    Escolhe a data mais próxima (ontem/hoje/amanhã) cujo weekday combina com day_code.
    Resolve virada de dia e aulas cruzando 00:00.
    """
    dc = (day_code or "").strip().upper()
    if dc not in DAY_TO_WD:
        return None
    target_wd = DAY_TO_WD[dc]
    candidates = [now.date() - timedelta(days=1), now.date(), now.date() + timedelta(days=1)]
    for d in candidates:
        if d.weekday() == target_wd:
            return d
    return now.date()


def _item_interval_debug(now: datetime, it: ClassItem) -> Tuple[Optional[Tuple[datetime, datetime]], str]:
    """
    Retorna ((start_dt, end_dt), reason)
    reason:
      - "ok"
      - "invalid_day"
      - "invalid_time"
    """
    base_date = _best_date_for_daycode(now, it.day)
    if base_date is None:
        return None, "invalid_day"

    t_start = _parse_time_obj(it.start)
    t_end = _parse_time_obj(it.end)
    if t_start is None or t_end is None:
        return None, "invalid_time"

    start_dt = datetime.combine(base_date, t_start)
    end_dt = datetime.combine(base_date, t_end)

    # Se cruza meia-noite, empurra o fim para o dia seguinte.
    if end_dt <= start_dt:
        end_dt = end_dt + timedelta(days=1)

    return (start_dt, end_dt), "ok"


# -------------------------
# Weather icon mapping
# -------------------------

def _img_path_try(base: str) -> Optional[str]:
    candidates = [
        os.path.join(GRAPHICS_DIR, base),
        os.path.join(GRAPHICS_DIR, base + ".png"),
        os.path.join(GRAPHICS_DIR, base + ".PNG"),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def map_symbol_to_icon(symbol_code: Optional[str]) -> str:
    if not symbol_code:
        return "clouds"

    s = symbol_code.lower()
    if "thunder" in s:
        return "storm"
    if "snow" in s:
        return "snowflake"
    if "rain" in s or "sleet" in s:
        return "rainy-day"
    if "partlycloudy" in s:
        return "cloudy"
    if "cloudy" in s:
        return "clouds"
    if "clearsky" in s or "fair" in s:
        return "sun"
    return "clouds"


# -------------------------
# UI Widgets
# -------------------------

class RoundedFrame(tk.Canvas):
    def __init__(self, master, radius=16, bg="#ffffff", border="#d0d7e2", **kwargs):
        super().__init__(master, highlightthickness=0, bd=0, bg=master["bg"], **kwargs)
        self.radius = radius
        self.fill = bg
        self.border = border
        self.bind("<Configure>", self._redraw)

    def _redraw(self, _evt=None):
        self.delete("all")
        w = self.winfo_width()
        h = self.winfo_height()
        r = min(self.radius, w // 2, h // 2)

        x1, y1, x2, y2 = 1, 1, w - 2, h - 2
        self._round_rect(x1, y1, x2, y2, r, fill=self.fill, outline=self.border)

    def _round_rect(self, x1, y1, x2, y2, r, fill, outline):
        points = [
            x1 + r, y1,
            x2 - r, y1,
            x2, y1,
            x2, y1 + r,
            x2, y2 - r,
            x2, y2,
            x2 - r, y2,
            x1 + r, y2,
            x1, y2,
            x1, y2 - r,
            x1, y1 + r,
            x1, y1
        ]
        self.create_polygon(points, smooth=True, fill=fill, outline=outline)


class ClassCard(tk.Frame):
    def __init__(self, master, is_day_theme: bool):
        super().__init__(master, bd=0, highlightthickness=0)
        self.is_day_theme = is_day_theme

        self.bg = "#ffffff" if is_day_theme else "#0f1a2a"
        self.border = "#d7dee8" if is_day_theme else "#20324d"
        self.text1 = "#0b2d4d" if is_day_theme else "#eaf2ff"
        self.text2 = "#5c6f86" if is_day_theme else "#b9c7dd"
        self.green = "#1aa56a"

        self.canvas = RoundedFrame(self, radius=14, bg=self.bg, border=self.border)
        self.canvas.pack(fill="both", expand=True)

        self.time_lbl = tk.Label(self.canvas, text="", font=("Segoe UI", 12, "bold"),
                                 fg=self.text1, bg=self.bg)
        self.title_lbl = tk.Label(self.canvas, text="", font=("Segoe UI", 16, "bold"),
                                  fg=self.text1, bg=self.bg, anchor="w", justify="left", wraplength=520)
        self.sub_lbl = tk.Label(self.canvas, text="", font=("Segoe UI", 11),
                                fg=self.text2, bg=self.bg, anchor="w")

        self.tag_lbl = tk.Label(self.canvas, text="MENOR", font=("Segoe UI", 9, "bold"),
                                fg="#134d2d" if is_day_theme else "#0b2a18",
                                bg="#dff4e8", padx=10, pady=3)

        self.bar_bg = tk.Frame(self.canvas, bg="#e9eef5" if is_day_theme else "#1d2b41", height=10)
        self.bar_fg = tk.Frame(self.bar_bg, bg=self.green, height=10)

        self.canvas.bind("<Configure>", self._place)

        self._is_minor = False
        self._progress = 0.0

    def _place(self, _evt=None):
        w = self.canvas.winfo_width()
        h = self.canvas.winfo_height()

        pad = 16
        self.canvas.create_window(pad, 10, anchor="nw", window=self.time_lbl, width=w - 2 * pad)
        self.canvas.create_window(pad, 34, anchor="nw", window=self.title_lbl, width=w - 2 * pad - 90)
        self.canvas.create_window(pad, 70, anchor="nw", window=self.sub_lbl, width=w - 2 * pad)

        if self._is_minor:
            self.canvas.create_window(w - pad, 34, anchor="ne", window=self.tag_lbl)

        self.canvas.create_window(pad, h - 18, anchor="sw", window=self.bar_bg, width=w - 2 * pad, height=10)
        self._update_bar()

    def set_data(self, start: str, end: str, title: str, teacher: str, location: str, tag: str,
                 progress: float):
        self.time_lbl.configure(text=f"{start}–{end}")
        self.title_lbl.configure(text=title.upper())
        sub = " | ".join([x for x in [teacher, location] if x])
        self.sub_lbl.configure(text=sub)

        self._is_minor = (tag or "").strip().upper() == "MENOR"
        self._progress = max(0.0, min(1.0, progress))
        self._place()

    def _update_bar(self):
        w = self.bar_bg.winfo_width()
        if w <= 1:
            self.after(30, self._update_bar)
            return
        fg_w = int(w * self._progress)
        self.bar_fg.place(x=0, y=0, width=fg_w, height=10)
        self.bar_bg.configure(height=10)


class SectionFrame(tk.Frame):
    def __init__(self, master, title: str, is_day_theme: bool):
        super().__init__(master, bd=0, highlightthickness=0)

        self.is_day_theme = is_day_theme
        self.bg = "#f5f7fb" if is_day_theme else "#0a1322"
        self.title_fg = "#1f7ab8" if is_day_theme else "#4aa3ff"

        self.configure(bg=self.bg)

        self.title_lbl = tk.Label(self, text=title, font=("Segoe UI", 18, "bold"),
                                  fg=self.title_fg, bg=self.bg, anchor="w")
        self.title_lbl.pack(fill="x", padx=10, pady=(0, 10))

        self.grid_frame = tk.Frame(self, bg=self.bg)
        self.grid_frame.pack(fill="both", expand=True)

        self.cards: List[ClassCard] = []

    def set_cards(self, card_data: List[Tuple[str, str, str, str, str, str, float]]):
        for c in self.cards:
            c.destroy()
        self.cards.clear()

        cols = 2
        for i, data in enumerate(card_data):
            r = i // cols
            c = i % cols
            card = ClassCard(self.grid_frame, is_day_theme=self.is_day_theme)
            card.set_data(*data)
            card.grid(row=r, column=c, padx=10, pady=10, sticky="nsew")
            self.grid_frame.grid_columnconfigure(c, weight=1, uniform="col")
            self.cards.append(card)

        for r in range((len(card_data) + cols - 1) // cols):
            self.grid_frame.grid_rowconfigure(r, weight=0)


class HoursCard(tk.Frame):
    def __init__(self, master, is_day_theme: bool):
        super().__init__(master, bd=0, highlightthickness=0)
        self.is_day_theme = is_day_theme
        self.bg = "#ffffff" if is_day_theme else "#0f1a2a"
        self.border = "#d7dee8" if is_day_theme else "#20324d"

        self.canvas = RoundedFrame(self, radius=18, bg=self.bg, border=self.border)
        self.canvas.pack(fill="both", expand=True)

        self.title = tk.Label(self.canvas, text="CLUBE", font=("Segoe UI", 16, "bold"),
                              fg="#ffffff", bg="#4e8f3b", padx=14, pady=8, anchor="w")
        self.status = tk.Label(self.canvas, text="ABERTO AGORA", font=("Segoe UI", 22, "bold"),
                               fg="#ffffff", bg="#63a43f", padx=14, pady=10, anchor="w")
        self.sub = tk.Label(self.canvas, text="FECHA ÀS 22H", font=("Segoe UI", 18, "bold"),
                            fg="#8a1f1f", bg="#f3e88f", padx=14, pady=10, anchor="w")
        self.lines = tk.Label(self.canvas, text="", font=("Segoe UI", 12, "bold"),
                              fg="#27364a" if is_day_theme else "#d8e6ff", bg=self.bg, justify="left")

        self.canvas.bind("<Configure>", self._place)

        self._mode = 0
        self._items = self._build_items()

    def _build_items(self):
        return [
            {
                "name": "CLUBE",
                "rules": {
                    "SEG": ("06:00", "22:00"),
                    "TER": ("06:00", "22:00"),
                    "QUA": ("06:00", "22:00"),
                    "QUI": ("06:00", "22:00"),
                    "SEX": ("06:00", "22:00"),
                    "SAB": ("08:00", "20:00"),
                    "DOM": ("09:00", "20:00"),
                },
                "lines": "SEG–SEX  6H–22H      SÁB 8H–20H\nDOM/FER  9H–20H",
            },
            {
                "name": "SECRETARIA",
                "rules": {
                    "SEG": ("08:00", "20:00"),
                    "TER": ("08:00", "20:00"),
                    "QUA": ("08:00", "20:00"),
                    "QUI": ("08:00", "20:00"),
                    "SEX": ("08:00", "20:00"),
                    "SAB": ("08:00", "20:00"),
                    "DOM": ("09:00", "20:00"),
                },
                "lines": "SEG–SEX  8H–20H      SÁB 8H–20H\nDOM/FER  9H–20H",
            },
            {
                "name": "ACADEMIA",
                # ✅ horários corrigidos:
                # SEG–QUI 06:00–21:30 | SEX 06:00–21:00 | SÁB 08:00–12:00 | DOM FECHADO
                "rules": {
                    "SEG": ("06:00", "21:30"),
                    "TER": ("06:00", "21:30"),
                    "QUA": ("06:00", "21:30"),
                    "QUI": ("06:00", "21:30"),
                    "SEX": ("06:00", "21:00"),
                    "SAB": ("08:00", "12:00"),
                    "DOM": (None, None),
                },
                "lines": "SEG–QUI  06:00–21:30\nSEX      06:00–21:00\nSÁB      08:00–12:00\nDOM      FECHADO",
            }
        ]

    def _place(self, _evt=None):
        w = self.canvas.winfo_width()
        h = self.canvas.winfo_height()

        self.canvas.create_window(20, 16, anchor="nw", window=self.title, width=w - 40, height=44)
        self.canvas.create_window(20, 70, anchor="nw", window=self.status, width=w - 40, height=66)
        self.canvas.create_window(20, 144, anchor="nw", window=self.sub, width=w - 40, height=54)
        self.canvas.create_window(20, 210, anchor="nw", window=self.lines, width=w - 40, height=h - 220)

    def tick_rotate(self):
        self._mode = (self._mode + 1) % len(self._items)
        self.update_view()

    def update_view(self):
        item = self._items[self._mode]
        self.title.configure(text=item["name"])

        day = today_3letters_noaccent()
        now = time.localtime()
        now_min = now.tm_hour * 60 + now.tm_min

        open_s, close_s = item["rules"].get(day, (None, None))
        open_m = parse_hhmm(open_s) if open_s else None
        close_m = parse_hhmm(close_s) if close_s else None

        if open_m is None or close_m is None:
            is_open = False
        else:
            is_open = (open_m <= now_min < close_m)

        if is_open:
            self.status.configure(text="ABERTO AGORA", bg="#63a43f")
            self.sub.configure(
                text=f"FECHA ÀS {fmt_hhmm(close_m)}",
                fg="#8a1f1f",
                bg="#f3e88f"
            )
        else:
            self.status.configure(text="FECHADO AGORA", bg="#3d556d")
            if open_m is not None:
                self.sub.configure(
                    text=f"ABRE ÀS {fmt_hhmm(open_m)}",
                    fg="#ffffff",
                    bg="#2b3f55"
                )
            else:
                self.sub.configure(text="SEM ATENDIMENTO", fg="#ffffff", bg="#2b3f55")

        self.lines.configure(text=item["lines"])


class WeatherCard(tk.Frame):
    def __init__(self, master, is_day_theme: bool):
        super().__init__(master, bd=0, highlightthickness=0)
        self.is_day_theme = is_day_theme

        self.bg = "#0f4f9e" if is_day_theme else "#0b2b5e"
        self.border = "#083b7a" if is_day_theme else "#0a2248"

        self.canvas = RoundedFrame(self, radius=18, bg=self.bg, border=self.border)
        self.canvas.pack(fill="both", expand=True)

        self.title = tk.Label(self.canvas, text="CLIMA  ALFENAS", font=("Segoe UI", 18, "bold"),
                              fg="#ffffff", bg=self.bg, anchor="w")
        self.temp = tk.Label(self.canvas, text="—°C", font=("Segoe UI", 38, "bold"),
                             fg="#ffffff", bg=self.bg, anchor="e")
        self.desc = tk.Label(self.canvas, text="—", font=("Segoe UI", 18, "bold"),
                             fg="#ffffff", bg=self.bg, anchor="e")
        self.nextline = tk.Label(self.canvas, text="Amanhã: —", font=("Segoe UI", 14),
                                 fg="#d6e6ff", bg=self.bg, anchor="e")
        self.status = tk.Label(self.canvas, text="", font=("Segoe UI", 10, "bold"),
                               fg="#d6e6ff", bg=self.bg, anchor="e")

        self.icon_img = None
        self.icon_lbl = tk.Label(self.canvas, image=None, bg=self.bg)

        self.canvas.bind("<Configure>", self._place)

    def _place(self, _evt=None):
        w = self.canvas.winfo_width()
        h = self.canvas.winfo_height()

        pad = 18
        self.canvas.create_window(pad, 14, anchor="nw", window=self.title, width=w - 2 * pad, height=34)

        self.canvas.create_window(pad + 70, h // 2 + 10, anchor="center", window=self.icon_lbl, width=140, height=140)

        self.canvas.create_window(w - pad, 62, anchor="ne", window=self.temp, width=w - 220, height=56)
        self.canvas.create_window(w - pad, 122, anchor="ne", window=self.desc, width=w - 220, height=30)
        self.canvas.create_window(w - pad, 154, anchor="ne", window=self.nextline, width=w - 220, height=26)
        self.canvas.create_window(w - pad, h - 12, anchor="se", window=self.status, width=w - 2 * pad, height=18)

    def set_weather(self, city: str, res: weather_mod.WeatherResult):
        if res.temp_c is not None:
            self.temp.configure(text=f"{res.temp_c}°C")
        else:
            self.temp.configure(text="—°C")

        desc = res.today_label.split(":", 1)[-1].strip() if ":" in res.today_label else res.today_label
        self.desc.configure(text=desc.upper())

        self.nextline.configure(text=res.tomorrow_label)

        # ✅ status sempre visível (ONLINE/OFFLINE/SEM DADOS)
        if res.ok:
            if res.cache_ts:
                hhmm = time.strftime("%H:%M", time.localtime(res.cache_ts))
            else:
                hhmm = "—"
            if res.source == "cache":
                self.status.configure(text=f"OFFLINE • cache {hhmm}")
            else:
                self.status.configure(text=f"ONLINE • atualizado {hhmm}")
        else:
            self.status.configure(text="SEM DADOS")

        icon_base = map_symbol_to_icon(res.symbol_code)
        p = _img_path_try(icon_base) or _img_path_try("clouds")
        if p:
            try:
                self.icon_img = tk.PhotoImage(file=p)
                if self.icon_img.width() > 200:
                    factor = max(1, self.icon_img.width() // 160)
                    self.icon_img = self.icon_img.subsample(factor, factor)
                self.icon_lbl.configure(image=self.icon_img)
            except Exception:
                self.icon_lbl.configure(image=None)


# -------------------------
# Main App
# -------------------------

class SALApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("SAL - SESI Agenda Live")
        self.attributes("-fullscreen", True)
        self.configure(bg="#f5f7fb")

        self.is_day_theme = theme_is_day()
        self._apply_theme()

        # header
        self.header = tk.Frame(self, bg=self.bg_header, height=90)
        self.header.pack(fill="x", side="top")

        self.header.grid_columnconfigure(0, weight=0)
        self.header.grid_columnconfigure(1, weight=1)
        self.header.grid_columnconfigure(2, weight=0)

        self.logo_img = None
        self.logo_lbl = tk.Label(self.header, bg=self.bg_header)
        self.logo_lbl.grid(row=0, column=0, rowspan=2, sticky="w", padx=18, pady=10)

        self.date_lbl = tk.Label(self.header, text="", font=("Segoe UI", 18, "bold"),
                                 fg=self.fg_primary, bg=self.bg_header)
        self.time_lbl = tk.Label(self.header, text="", font=("Segoe UI", 28, "bold"),
                                 fg=self.fg_primary, bg=self.bg_header)

        self.date_lbl.grid(row=0, column=1, sticky="w", padx=10, pady=(18, 0))
        self.time_lbl.grid(row=0, column=1, sticky="w", padx=160, pady=(8, 0))

        self.div = tk.Frame(self, bg=self.border_soft, height=2)
        self.div.pack(fill="x")

        self.main = tk.Frame(self, bg=self.bg_root)
        self.main.pack(fill="both", expand=True)

        self.main.grid_columnconfigure(0, weight=3, uniform="main")
        self.main.grid_columnconfigure(1, weight=2, uniform="main")
        self.main.grid_rowconfigure(0, weight=1)
        self.main.grid_rowconfigure(1, weight=1)

        self.agora = SectionFrame(self.main, "AGORA", is_day_theme=self.is_day_theme)
        self.prox = SectionFrame(self.main, "PRÓXIMAS", is_day_theme=self.is_day_theme)

        self.agora.grid(row=0, column=0, sticky="nsew", padx=(16, 8), pady=(16, 8))
        self.prox.grid(row=1, column=0, sticky="nsew", padx=(16, 8), pady=(8, 16))

        self.hours_card = HoursCard(self.main, is_day_theme=self.is_day_theme)
        self.weather_card = WeatherCard(self.main, is_day_theme=self.is_day_theme)

        self.hours_card.grid(row=0, column=1, sticky="nsew", padx=(8, 16), pady=(16, 8))
        self.weather_card.grid(row=1, column=1, sticky="nsew", padx=(8, 16), pady=(8, 16))

        self.all_items: List[ClassItem] = []
        self.last_excel_mtime: Optional[float] = None

        # ✅ backend bônus: clima não-bloqueante (thread)
        self.weather_last_fetch = 0.0
        self.weather_res: Optional[weather_mod.WeatherResult] = None
        self._weather_lock = threading.Lock()
        self._weather_inflight = False

        self._last_zero_agenda_log_ts = 0.0
        self._last_housekeeping_ts = 0.0

        self.bind("<Escape>", lambda e: self.destroy())

        # boot logs + housekeeping
        _rotate_logs_if_needed(logger=log)

        log(f"[BOOT] APP_DIR={APP_DIR}")
        log(f"[BOOT] DATA_DIR={DATA_DIR}")
        log(f"[BOOT] EXCEL_PATH={EXCEL_PATH} exists={os.path.exists(EXCEL_PATH)}")
        log(f"[BOOT] GRAPHICS_DIR={GRAPHICS_DIR} exists={os.path.exists(GRAPHICS_DIR)}")
        log(f"[BOOT] LOG_PATH={LOG_PATH}")

        # ✅ chama housekeeping do clima (limpa cache_old etc.)
        try:
            weather_mod.housekeeping(app_dir=APP_DIR, logger=log)
        except Exception as e:
            log(f"[WEATHER] Housekeeping error {type(e).__name__}: {e}")

        self._refresh_logo()
        self._reload_excel_if_needed(force=True)
        self._tick()

        self.after(9000, self._rotate_hours)

    def _apply_theme(self):
        self.is_day_theme = theme_is_day()
        self.bg_root = "#f5f7fb" if self.is_day_theme else "#06101f"
        self.bg_header = "#f5f7fb" if self.is_day_theme else "#06101f"
        self.fg_primary = "#0b2d4d" if self.is_day_theme else "#eaf2ff"
        self.border_soft = "#cfd8e5" if self.is_day_theme else "#132744"
        self.configure(bg=self.bg_root)

    def _refresh_logo(self):
        base = "logo_day" if self.is_day_theme else "logo_night"
        p = _img_path_try(base)
        if not p:
            p = _img_path_try("logo_sesi") or _img_path_try("logo_day") or _img_path_try("logo_night")
        if p:
            try:
                img = tk.PhotoImage(file=p)
                if img.width() > 280:
                    factor = max(1, img.width() // 260)
                    img = img.subsample(factor, factor)
                self.logo_img = img
                self.logo_lbl.configure(image=self.logo_img)
            except Exception:
                self.logo_lbl.configure(image=None)

    def _rotate_hours(self):
        try:
            self.hours_card.tick_rotate()
        finally:
            self.after(9000, self._rotate_hours)

    def _reload_excel_if_needed(self, force: bool = False):
        try:
            mtime = os.path.getmtime(EXCEL_PATH)
            if force or self.last_excel_mtime is None or mtime != self.last_excel_mtime:
                items = load_classes_from_excel(EXCEL_PATH)
                self.all_items = items
                self.last_excel_mtime = mtime
                log(f"[XLSX] Excel carregado: {len(self.all_items)} itens. mtime={mtime}")
        except Exception as e:
            log(f"[XLSX] Falha ao carregar Excel: {type(e).__name__}: {e}")

    def _compute_now_next(self) -> Tuple[List[Tuple], List[Tuple]]:
        """
        AGORA/PRÓXIMAS robusto com virada de dia e cruzamento da meia-noite.
        Também gera logs diagnósticos quando 0 itens visíveis.
        """
        now_dt = datetime.now()
        window_end = now_dt + timedelta(minutes=120)

        now_list = []
        next_list = []

        discard_day = 0
        discard_time = 0
        discard_other = 0

        for it in self.all_items:
            interval, reason = _item_interval_debug(now_dt, it)
            if not interval:
                if reason == "invalid_day":
                    discard_day += 1
                elif reason == "invalid_time":
                    discard_time += 1
                else:
                    discard_other += 1
                continue

            start_dt, end_dt = interval

            if start_dt <= now_dt < end_dt:
                total = (end_dt - start_dt).total_seconds()
                done = (now_dt - start_dt).total_seconds()
                progress = 0.0 if total <= 0 else max(0.0, min(1.0, done / total))
                now_list.append((it.start, it.end, it.activity, it.teacher, it.location, it.tag, progress, end_dt, start_dt))
            elif now_dt <= start_dt <= window_end:
                next_list.append((it.start, it.end, it.activity, it.teacher, it.location, it.tag, 0.0, end_dt, start_dt))

        # sort: AGORA por término mais cedo; PRÓXIMAS por início
        now_list.sort(key=lambda t: t[7])   # end_dt
        next_list.sort(key=lambda t: t[8])  # start_dt

        # formato da UI
        now_cards = [(a, b, c, d, e, f, g) for (a, b, c, d, e, f, g, _end, _start) in now_list]
        next_cards = [(a, b, c, d, e, f, g) for (a, b, c, d, e, f, g, _end, _start) in next_list]

        # log quando 0 itens visíveis (rate limit ~60s para não poluir)
        if self.all_items and (len(now_cards) == 0 and len(next_cards) == 0):
            if time.time() - self._last_zero_agenda_log_ts > 60:
                self._last_zero_agenda_log_ts = time.time()
                sample = self.all_items[:4]
                log(
                    "[AGENDA] 0 em AGORA/PRÓXIMAS | "
                    f"now={now_dt.strftime('%Y-%m-%d %H:%M:%S')} window_end={window_end.strftime('%Y-%m-%d %H:%M:%S')} "
                    f"today_code={today_3letters_noaccent()} "
                    f"discard_day={discard_day} discard_time={discard_time} discard_other={discard_other} "
                    f"sample_items={sample}"
                )

        return now_cards[:6], next_cards[:6]

    # ✅ backend bônus: fetch do clima em background para não travar a UI
    def _weather_worker(self):
        try:
            res = weather_mod.get_weather(
                city_label="Alfenas",
                lat=-21.4267,
                lon=-45.9470,
                app_dir=APP_DIR,
                user_agent="SAL-SESIAgendaLive/2.0 (contact: local)",
                logger=log,
            )
            with self._weather_lock:
                self.weather_res = res
                self.weather_last_fetch = time.time()
        except Exception as e:
            log(f"[WEATHER] Worker error {type(e).__name__}: {e}")
        finally:
            with self._weather_lock:
                self._weather_inflight = False

    def _tick_weather(self):
        # refresh a cada 10 minutos (ou se nunca buscou)
        if (time.time() - self.weather_last_fetch) < 600 and self.weather_res is not None:
            return

        with self._weather_lock:
            if self._weather_inflight:
                return
            self._weather_inflight = True

        th = threading.Thread(target=self._weather_worker, daemon=True)
        th.start()

    def _tick_housekeeping(self):
        # housekeeping leve 1x por dia (não precisa ser mais que isso)
        if time.time() - self._last_housekeeping_ts < 86400:
            return
        self._last_housekeeping_ts = time.time()
        try:
            _rotate_logs_if_needed(logger=log)
            weather_mod.housekeeping(app_dir=APP_DIR, logger=log)
        except Exception as e:
            log(f"[HK] error {type(e).__name__}: {e}")

    def _tick(self):
        try:
            new_theme = theme_is_day()
            if new_theme != self.is_day_theme:
                self._apply_theme()
                self.header.configure(bg=self.bg_header)
                self.div.configure(bg=self.border_soft)
                self.main.configure(bg=self.bg_root)
                self.date_lbl.configure(bg=self.bg_header, fg=self.fg_primary)
                self.time_lbl.configure(bg=self.bg_header, fg=self.fg_primary)
                self.logo_lbl.configure(bg=self.bg_header)

                self.agora.destroy()
                self.prox.destroy()
                self.hours_card.destroy()
                self.weather_card.destroy()

                self.agora = SectionFrame(self.main, "AGORA", is_day_theme=self.is_day_theme)
                self.prox = SectionFrame(self.main, "PRÓXIMAS", is_day_theme=self.is_day_theme)
                self.agora.grid(row=0, column=0, sticky="nsew", padx=(16, 8), pady=(16, 8))
                self.prox.grid(row=1, column=0, sticky="nsew", padx=(16, 8), pady=(8, 16))

                self.hours_card = HoursCard(self.main, is_day_theme=self.is_day_theme)
                self.weather_card = WeatherCard(self.main, is_day_theme=self.is_day_theme)
                self.hours_card.grid(row=0, column=1, sticky="nsew", padx=(8, 16), pady=(16, 8))
                self.weather_card.grid(row=1, column=1, sticky="nsew", padx=(8, 16), pady=(8, 16))

                self._refresh_logo()

            d, t = date_time_strings()
            self.date_lbl.configure(text=d)
            self.time_lbl.configure(text=t)

            self.hours_card.update_view()

            self._reload_excel_if_needed(force=False)

            now_cards, next_cards = self._compute_now_next()
            self.agora.set_cards(now_cards)
            self.prox.set_cards(next_cards)

            self._tick_weather()

            # aplica na UI o último resultado disponível (sem travar)
            with self._weather_lock:
                res = self.weather_res
            if res:
                self.weather_card.set_weather("Alfenas", res)

            self._tick_housekeeping()

        except Exception:
            log("Tick error:\n" + traceback.format_exc())

        self.after(1000, self._tick)


if __name__ == "__main__":
    try:
        SALApp().mainloop()
    except Exception:
        log("Fatal error:\n" + traceback.format_exc())
