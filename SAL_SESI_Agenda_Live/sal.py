# sal.py
# SAL - SESI Agenda Live (v2 UI - Layout A)
# Layout: Header (Logo + Date/Time/Weekday + Hours/Weather). Body: 2 big columns (AGORA | PRÓXIMAS).
# Assets: graphics/logo_day(.png), graphics/logo_night(.png), graphics/*.png (icons)
# Logs (gravável): %LOCALAPPDATA%/SAL_SESI_Agenda_Live/logs/sal.log
# Cache clima (gravável): %LOCALAPPDATA%/SAL_SESI_Agenda_Live/package/weather_cache.json (+ package/cache_old/*.json)
# Excel: grade.xlsx (no APP_DIR)

from __future__ import annotations

import os
import sys
import time
import traceback
import threading
from dataclasses import dataclass
from typing import List, Optional, Tuple, Any

import tkinter as tk
import tkinter.font as tkfont  # UI: auto-fit fonts
from openpyxl import load_workbook

import weather as weather_mod

from datetime import datetime, timedelta, date as dt_date, time as dt_time


SAL_UI_BUILD = "UI_BUILD_2026-02-11A"


# -------------------------
# Paths / Logging
# -------------------------

def app_dir() -> str:
    # Works for python and PyInstaller
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def data_dir(app_name: str = "SAL_SESI_Agenda_Live") -> str:
    """
    Diretório gravável por usuário.
    - Em PCs corporativos, evita problemas quando o app roda em Program Files.
    - Se LOCALAPPDATA não existir, cai no APP_DIR (melhor esforço).
    """
    base = os.environ.get("LOCALAPPDATA")
    if not base:
        p = os.path.join(app_dir(), "_data")
        os.makedirs(p, exist_ok=True)
        return p
    p = os.path.join(base, app_name)
    os.makedirs(p, exist_ok=True)
    return p


APP_DIR = app_dir()
DATA_DIR = data_dir()

GRAPHICS_DIR = os.path.join(APP_DIR, "graphics")

# logs em local gravável
LOGS_DIR = os.path.join(DATA_DIR, "logs")
os.makedirs(LOGS_DIR, exist_ok=True)
LOG_PATH = os.path.join(LOGS_DIR, "sal.log")

# rotação/limpeza de logs
LOG_ARCHIVE_DIR = os.path.join(LOGS_DIR, "archive")
os.makedirs(LOG_ARCHIVE_DIR, exist_ok=True)

LOG_ROTATE_MAX_BYTES = 2 * 1024 * 1024   # 2MB
LOG_ARCHIVE_KEEP = 6                     # mantém os últimos N logs arquivados
LOG_ARCHIVE_MAX_AGE_DAYS = 30            # e/ou apaga logs muito antigos


def _safe_unlink(path: str) -> None:
    try:
        os.remove(path)
    except Exception:
        pass


def _cleanup_log_archive() -> None:
    try:
        files = []
        now = time.time()
        max_age = LOG_ARCHIVE_MAX_AGE_DAYS * 86400

        for name in os.listdir(LOG_ARCHIVE_DIR):
            p = os.path.join(LOG_ARCHIVE_DIR, name)
            if not os.path.isfile(p):
                continue
            try:
                st = os.stat(p)
            except Exception:
                continue

            if max_age > 0 and (now - st.st_mtime) > max_age:
                _safe_unlink(p)
                continue

            files.append((st.st_mtime, p))

        files.sort(reverse=True)  # mais novo primeiro
        for _mtime, p in files[LOG_ARCHIVE_KEEP:]:
            _safe_unlink(p)

    except Exception:
        pass


def _rotate_logs_if_needed(logger=None) -> None:
    """
    Rotaciona sal.log quando cresce demais.
    - Move para logs/archive/sal_YYYYMMDD_HHMMSS.log
    - Mantém no máximo LOG_ARCHIVE_KEEP arquivos e remove os muito antigos
    """
    try:
        if not os.path.exists(LOG_PATH):
            return
        size = os.path.getsize(LOG_PATH)
        if size < LOG_ROTATE_MAX_BYTES:
            _cleanup_log_archive()
            return

        ts = time.strftime("%Y%m%d_%H%M%S")
        archived = os.path.join(LOG_ARCHIVE_DIR, f"sal_{ts}.log")
        try:
            os.replace(LOG_PATH, archived)
        except Exception:
            with open(LOG_PATH, "rb") as src, open(archived, "wb") as dst:
                dst.write(src.read())
            with open(LOG_PATH, "w", encoding="utf-8"):
                pass

        if logger:
            logger(f"[LOG] Rotated sal.log -> {archived}")

        _cleanup_log_archive()

    except Exception as e:
        if logger:
            logger(f"[LOG] Rotate error {type(e).__name__}: {e}")


def log(msg: str) -> None:
    try:
        _rotate_logs_if_needed()
        ts = time.strftime("%Y-%m-%d %H:%M:%S")
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass


# -------------------------
# Debounce helper (anti-flicker)
# -------------------------

def _debounce_after(widget: tk.Misc, attr_name: str, delay_ms: int, fn):
    """
    Debounce simples por widget:
    - cancela o after anterior guardado em widget.<attr_name>
    - agenda um novo após delay_ms
    """
    job = getattr(widget, attr_name, None)
    if job is not None:
        try:
            widget.after_cancel(job)
        except Exception:
            pass
    new_job = widget.after(delay_ms, fn)
    setattr(widget, attr_name, new_job)


# -------------------------
# Data model
# -------------------------

@dataclass
class ClassItem:
    day: str          # SEG TER QUA QUI SEX SAB DOM
    start: str        # HH:MM
    end: str          # HH:MM
    activity: str
    teacher: str
    location: str
    tag: str          # MENOR/GERAL/...


# -------------------------
# Time parsing helpers (tolerante)
# -------------------------

def parse_hhmm(s: str) -> Optional[int]:
    """
    Aceita:
    - "HH:MM"
    - "HH:MM:SS"
    """
    try:
        s = str(s).strip()
        if not s:
            return None
        parts = s.split(":")
        if len(parts) < 2:
            return None
        hh = int(parts[0])
        mm = int(parts[1])
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return hh * 60 + mm
        return None
    except Exception:
        return None


def _parse_time_obj(s: str) -> Optional[dt_time]:
    m = parse_hhmm(s)
    if m is None:
        return None
    return dt_time(hour=m // 60, minute=m % 60)


def today_3letters_noaccent() -> str:
    wd = time.localtime().tm_wday  # Monday=0
    return ["SEG", "TER", "QUA", "QUI", "SEX", "SAB", "DOM"][wd]


def weekday_full_noaccent() -> str:
    wd = time.localtime().tm_wday
    return [
        "SEGUNDA-FEIRA",
        "TERCA-FEIRA",
        "QUARTA-FEIRA",
        "QUINTA-FEIRA",
        "SEXTA-FEIRA",
        "SABADO",
        "DOMINGO",
    ][wd]


def date_time_strings() -> Tuple[str, str, str]:
    lt = time.localtime()
    date_s = time.strftime("%d/%m/%Y", lt)
    time_s = time.strftime("%H:%M:%S", lt)  # com segundos (relógio do cabeçalho)
    wd_s = weekday_full_noaccent()
    return date_s, time_s, wd_s


def theme_is_day() -> bool:
    h = time.localtime().tm_hour
    return 6 <= h <= 17


def fmt_hhmm(mins: Optional[int]) -> str:
    if mins is None:
        return "—"
    hh = mins // 60
    mm = mins % 60
    if mm == 0:
        return f"{hh:02d}H"
    return f"{hh:02d}:{mm:02d}"


# -------------------------
# Excel read
# -------------------------

EXCEL_PATH = os.path.join(APP_DIR, "grade.xlsx")
SHEET_NAME = "SAL"
EXPECTED_HEADERS = ["DIA", "INICIO", "FIM", "ATIVIDADE", "PROFESSOR", "LOCAL", "TAG"]


def _normalize_header(v: object) -> str:
    return str(v or "").strip().upper()


def load_classes_from_excel(path: str) -> List[ClassItem]:
    wb = load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Aba '{SHEET_NAME}' não encontrada.")
    ws = wb[SHEET_NAME]

    header_row = None
    headers: List[str] = []
    for r in range(1, 6):
        vals = [_normalize_header(ws.cell(row=r, column=c).value) for c in range(1, 30)]
        if "DIA" in vals and "INICIO" in vals and "FIM" in vals:
            header_row = r
            headers = vals
            break
    if header_row is None:
        header_row = 1
        headers = [_normalize_header(ws.cell(row=1, column=c).value) for c in range(1, 30)]

    def col_idx(name: str) -> Optional[int]:
        name = name.upper()
        if name in headers:
            return headers.index(name) + 1
        return None

    c_dia = col_idx("DIA")
    c_ini = col_idx("INICIO")
    c_fim = col_idx("FIM")
    c_ativ = col_idx("ATIVIDADE")
    c_prof = col_idx("PROFESSOR")
    c_loc = col_idx("LOCAL")
    c_tag = col_idx("TAG")

    if not (c_dia and c_ini and c_fim and c_ativ):
        raise RuntimeError("Cabeçalhos necessários não encontrados na aba SAL.")

    items: List[ClassItem] = []
    for r in range(header_row + 1, ws.max_row + 1):
        dia = str(ws.cell(row=r, column=c_dia).value or "").strip().upper()
        ini = str(ws.cell(row=r, column=c_ini).value or "").strip()
        fim = str(ws.cell(row=r, column=c_fim).value or "").strip()
        ativ = str(ws.cell(row=r, column=c_ativ).value or "").strip()

        if not dia or not ini or not fim or not ativ:
            continue

        prof = str(ws.cell(row=r, column=c_prof).value or "").strip() if c_prof else ""
        loc = str(ws.cell(row=r, column=c_loc).value or "").strip() if c_loc else ""
        tag = str(ws.cell(row=r, column=c_tag).value or "").strip().upper() if c_tag else ""

        items.append(ClassItem(dia, ini, fim, ativ, prof, loc, tag))

    return items


# -------------------------
# Day mapping + midnight-safe datetime placement
# -------------------------

DAY_TO_WD = {"SEG": 0, "TER": 1, "QUA": 2, "QUI": 3, "SEX": 4, "SAB": 5, "DOM": 6}


def _best_date_for_daycode(now: datetime, day_code: str) -> Optional[dt_date]:
    dc = (day_code or "").strip().upper()
    if dc not in DAY_TO_WD:
        return None
    target_wd = DAY_TO_WD[dc]
    candidates = [now.date() - timedelta(days=1), now.date(), now.date() + timedelta(days=1)]
    for d in candidates:
        if d.weekday() == target_wd:
            return d
    return now.date()


def _item_interval_debug(now: datetime, it: ClassItem) -> Tuple[Optional[Tuple[datetime, datetime]], str]:
    base_date = _best_date_for_daycode(now, it.day)
    if base_date is None:
        return None, "invalid_day"

    t_start = _parse_time_obj(it.start)
    t_end = _parse_time_obj(it.end)
    if t_start is None or t_end is None:
        return None, "invalid_time"

    start_dt = datetime.combine(base_date, t_start)
    end_dt = datetime.combine(base_date, t_end)

    if end_dt <= start_dt:
        end_dt = end_dt + timedelta(days=1)

    return (start_dt, end_dt), "ok"


# -------------------------
# Weather icon mapping
# -------------------------

def _img_path_try(base: str) -> Optional[str]:
    candidates = [
        os.path.join(GRAPHICS_DIR, base),
        os.path.join(GRAPHICS_DIR, base + ".png"),
        os.path.join(GRAPHICS_DIR, base + ".PNG"),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def map_symbol_to_icon(symbol_code: Optional[str]) -> str:
    if not symbol_code:
        return "clouds"

    s = str(symbol_code).lower()
    if "thunder" in s:
        return "storm"
    if "snow" in s:
        return "snowflake"
    if "rain" in s or "sleet" in s:
        return "rainy-day"
    if "partlycloudy" in s:
        return "cloudy"
    if "cloudy" in s:
        return "clouds"
    if "clearsky" in s or "fair" in s:
        return "sun"
    return "clouds"


def _get_any(obj: Any, *keys: str, default=None):
    """Tenta ler obj.key, obj['key'] para múltiplas chaves."""
    for k in keys:
        # atributo
        try:
            if hasattr(obj, k):
                v = getattr(obj, k)
                if v is not None:
                    return v
        except Exception:
            pass
        # dict-like
        try:
            if isinstance(obj, dict) and k in obj:
                v = obj.get(k)
                if v is not None:
                    return v
        except Exception:
            pass
    return default


# -------------------------
# UI Widgets
# -------------------------

class RoundedFrame(tk.Canvas):
    """
    Canvas com retângulo arredondado + sombra leve (apenas estética).
    Anti-flicker: redraw com debounce e só quando tamanho muda.
    """
    def __init__(self, master, radius=16, bg="#ffffff", border="#d0d7e2",
                 shadow=False, shadow_offset=(3, 4), shadow_stipple="gray25", **kwargs):
        super().__init__(master, highlightthickness=0, bd=0, bg=master["bg"], **kwargs)
        self.radius = radius
        self.fill = bg
        self.border = border
        self.shadow = shadow
        self.shadow_offset = shadow_offset
        self.shadow_stipple = shadow_stipple

        self._last_wh = (0, 0)
        self._redraw_job = None
        self.bind("<Configure>", self._on_cfg)

    def _on_cfg(self, _evt=None):
        w = self.winfo_width()
        h = self.winfo_height()
        if (w, h) == self._last_wh:
            return
        self._last_wh = (w, h)
        _debounce_after(self, "_redraw_job", 33, self._redraw)

    def _redraw(self, _evt=None):
        self.delete("all")
        w = self.winfo_width()
        h = self.winfo_height()
        if w <= 2 or h <= 2:
            return

        r = min(self.radius, w // 2, h // 2)
        x1, y1, x2, y2 = 1, 1, w - 2, h - 2

        if self.shadow:
            dx, dy = self.shadow_offset
            self._round_rect(
                x1 + dx, y1 + dy, x2 + dx, y2 + dy, r,
                fill="#000000", outline="#000000",
                tags=("shadow",)
            )
            self.itemconfig("shadow", stipple=self.shadow_stipple)

        self._round_rect(x1, y1, x2, y2, r, fill=self.fill, outline=self.border, tags=("main",))

    def _round_rect(self, x1, y1, x2, y2, r, fill, outline, tags=()):
        points = [
            x1 + r, y1,
            x2 - r, y1,
            x2, y1,
            x2, y1 + r,
            x2, y2 - r,
            x2, y2,
            x2 - r, y2,
            x1 + r, y2,
            x1, y2,
            x1, y2 - r,
            x1, y1 + r,
            x1, y1
        ]
        self.create_polygon(points, smooth=True, fill=fill, outline=outline, tags=tags)


class ClassCard(tk.Frame):
    """
    Card compacto:
    - altura fixa (densidade)
    - progress bar “colada” no conteúdo
    - pill MENOR consistente e alinhada
    Anti-flicker: place com debounce e só quando tamanho muda.
    """
    def __init__(self, master, is_day_theme: bool):
        super().__init__(master, bd=0, highlightthickness=0)
        self.is_day_theme = is_day_theme

        self.CARD_H = 140
        self.configure(height=self.CARD_H)
        self.pack_propagate(False)
        self.grid_propagate(False)

        self.bg = "#ffffff" if is_day_theme else "#0f1a2a"
        self.border = "#d7dee8" if is_day_theme else "#20324d"
        self.text1 = "#0b2d4d" if is_day_theme else "#eaf2ff"
        self.text2 = "#5c6f86" if is_day_theme else "#b9c7dd"
        self.green = "#1aa56a"

        shadow_stipple = "gray25" if is_day_theme else "gray12"

        self.canvas = RoundedFrame(self, radius=14, bg=self.bg, border=self.border,
                                   shadow=True, shadow_offset=(2, 3), shadow_stipple=shadow_stipple)
        self.canvas.pack(fill="both", expand=True)

        self.time_lbl = tk.Label(self.canvas, text="", font=("Segoe UI", 11, "bold"),
                                 fg=self.text1, bg=self.bg, anchor="e")
        self.title_lbl = tk.Label(self.canvas, text="", font=("Segoe UI", 15, "bold"),
                                  fg=self.text1, bg=self.bg, anchor="w",
                                  justify="left", wraplength=520)
        self.sub_lbl = tk.Label(self.canvas, text="", font=("Segoe UI", 10),
                                fg=self.text2, bg=self.bg, anchor="w")

        self.tag_lbl = tk.Label(
            self.canvas, text="MENOR",
            font=("Segoe UI", 9, "bold"),
            fg="#0b2a18", bg="#dff4e8",
            padx=10, pady=2
        )
        self.tag_lbl.configure(highlightthickness=1, highlightbackground="#8fd3b2",
                               highlightcolor="#8fd3b2", bd=0)

        self.bar_bg = tk.Frame(self.canvas, bg="#e9eef5" if is_day_theme else "#1d2b41", height=10)
        self.bar_fg = tk.Frame(self.bar_bg, bg=self.green, height=10)

        self._is_minor = False
        self._progress = 0.0

        self._place_job = None
        self._last_wh = (0, 0)
        self.canvas.bind("<Configure>", self._on_cfg)

    def _on_cfg(self, _evt=None):
        w = self.canvas.winfo_width()
        h = self.canvas.winfo_height()
        if (w, h) == self._last_wh:
            return
        self._last_wh = (w, h)
        _debounce_after(self.canvas, "_place_job", 33, self._place)

    def _place(self, _evt=None):
        self.canvas.delete("win")
        w = self.canvas.winfo_width()
        h = self.canvas.winfo_height()
        if w <= 10 or h <= 10:
            return

        pad_x = 14
        top = 12

        title_w = max(10, w - 2 * pad_x - 120)

        self.canvas.create_window(pad_x, top, anchor="nw", window=self.title_lbl,
                                  width=title_w, height=24, tags=("win",))
        self.canvas.create_window(w - pad_x, top, anchor="ne", window=self.time_lbl,
                                  width=110, height=24, tags=("win",))

        self.canvas.create_window(pad_x, top + 30, anchor="nw", window=self.sub_lbl,
                                  width=w - 2 * pad_x, height=18, tags=("win",))

        if self._is_minor:
            self.canvas.create_window(w - pad_x, top + 52, anchor="ne", window=self.tag_lbl, tags=("win",))

        bar_y = h - 16
        self.canvas.create_window(pad_x, bar_y, anchor="sw", window=self.bar_bg,
                                  width=w - 2 * pad_x, height=10, tags=("win",))
        self._update_bar()

    def set_data(self, start: str, end: str, title: str, teacher: str, location: str, tag: str,
                 progress: float):
        start_s = str(start).strip()
        end_s = str(end).strip()
        if len(start_s.split(":")) >= 2:
            start_s = ":".join(start_s.split(":")[:2])
        if len(end_s.split(":")) >= 2:
            end_s = ":".join(end_s.split(":")[:2])

        self.time_lbl.configure(text=f"{start_s}–{end_s}")
        self.title_lbl.configure(text=str(title).upper())

        sub = " | ".join([x for x in [teacher, location] if x])
        self.sub_lbl.configure(text=sub)

        self._is_minor = (tag or "").strip().upper() == "MENOR"
        self._progress = max(0.0, min(1.0, progress))
        self._place()

    def _update_bar(self):
        w = self.bar_bg.winfo_width()
        if w <= 1:
            self.after(20, self._update_bar)
            return
        fg_w = int(w * self._progress)
        self.bar_fg.place(x=0, y=0, width=fg_w, height=10)
        self.bar_bg.configure(height=10)


class SectionFrame(tk.Frame):
    """
    Seção:
    - painel com relevo (borda + sombra)
    - grid 2 colunas
    - cards FIXOS (pool)
    Anti-flicker: inner placement com debounce.
    """
    def __init__(self, master, title: str, is_day_theme: bool):
        super().__init__(master, bd=0, highlightthickness=0)

        self.is_day_theme = is_day_theme
        self.bg = "#f5f7fb" if is_day_theme else "#06101f"
        self.panel_bg = "#f5f7fb" if is_day_theme else "#06101f"
        self.title_fg = "#1f7ab8" if is_day_theme else "#4aa3ff"
        self.line_fg = "#d5deea" if is_day_theme else "#132744"
        self.panel_border = "#cfd8e5" if is_day_theme else "#112645"

        self.configure(bg=self.bg)

        shadow_stipple = "gray18" if is_day_theme else "gray10"

        self.panel = RoundedFrame(
            self, radius=18, bg=self.panel_bg, border=self.panel_border,
            shadow=True, shadow_offset=(3, 4), shadow_stipple=shadow_stipple
        )
        self.panel.pack(fill="both", expand=True)

        self.inner = tk.Frame(self.panel, bg=self.panel_bg)
        self.inner.configure(
            highlightthickness=1,
            highlightbackground=("#d0d7e2" if is_day_theme else "#0c2038"),
            bd=0
        )

        self._inner_job = None
        self._last_wh = (0, 0)
        self.panel.bind("<Configure>", self._on_panel_cfg)

        header = tk.Frame(self.inner, bg=self.panel_bg)
        header.pack(fill="x", padx=14, pady=(12, 10))

        self.title_lbl = tk.Label(
            header, text=title, font=("Segoe UI", 18, "bold"),
            fg=self.title_fg, bg=self.panel_bg, anchor="w"
        )
        self.title_lbl.pack(side="left")

        self.line = tk.Frame(header, bg=self.line_fg, height=2)
        self.line.pack(side="left", fill="x", expand=True, padx=(12, 0), pady=12)

        self.grid_frame = tk.Frame(self.inner, bg=self.panel_bg)
        self.grid_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.cols = 2
        self.pad = 10
        self.max_cards = 6

        self.cards: List[ClassCard] = []

        for c in range(self.cols):
            self.grid_frame.grid_columnconfigure(c, weight=1, uniform="col")

        for i in range(self.max_cards):
            r = i // self.cols
            c = i % self.cols
            card = ClassCard(self.grid_frame, is_day_theme=self.is_day_theme)
            card.grid(row=r, column=c, padx=self.pad, pady=self.pad, sticky="ew")
            self.cards.append(card)

        spacer_row = (self.max_cards + self.cols - 1) // self.cols
        self._spacer = tk.Frame(self.grid_frame, bg=self.panel_bg)
        self._spacer.grid(row=spacer_row, column=0, columnspan=self.cols, sticky="nsew")
        self.grid_frame.grid_rowconfigure(spacer_row, weight=1)

        for card in self.cards:
            card.grid_remove()

    def _on_panel_cfg(self, _evt=None):
        w = self.panel.winfo_width()
        h = self.panel.winfo_height()
        if (w, h) == self._last_wh:
            return
        self._last_wh = (w, h)
        _debounce_after(self.panel, "_inner_job", 33, self._place_inner)

    def _place_inner(self):
        w = self.panel.winfo_width()
        h = self.panel.winfo_height()
        if w <= 10 or h <= 10:
            return
        pad = 6
        self.panel.delete("inner")
        self.panel.create_window(
            pad, pad, anchor="nw", window=self.inner,
            width=max(10, w - 2 * pad),
            height=max(10, h - 2 * pad),
            tags=("inner",)
        )

    def set_cards(self, card_data: List[Tuple[str, str, str, str, str, str, float]]):
        n = min(len(card_data), self.max_cards)

        for i in range(n):
            self.cards[i].set_data(*card_data[i])
            self.cards[i].grid()

        for i in range(n, self.max_cards):
            self.cards[i].grid_remove()


class HoursCard(tk.Frame):
    """
    Card duplo (interligado) com auto-fit.
    Anti-flicker: <Configure> com debounce e só quando tamanho muda.
    """
    def __init__(self, master, is_day_theme: bool):
        super().__init__(master, bd=0, highlightthickness=0)
        self.is_day_theme = is_day_theme

        try:
            self.configure(bg=master["bg"])
        except Exception:
            pass

        self.bg = "#ffffff" if is_day_theme else "#0f1a2a"
        self.border = "#d7dee8" if is_day_theme else "#20324d"

        shadow_stipple = "gray18" if is_day_theme else "gray10"

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=5, uniform="hc")
        self.grid_columnconfigure(1, weight=0)
        self.grid_columnconfigure(2, weight=2, uniform="hc")

        self.left = RoundedFrame(
            self, radius=18, bg=self.bg, border=self.border,
            shadow=True, shadow_offset=(3, 4), shadow_stipple=shadow_stipple
        )
        self.left.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=0)

        self._connector = tk.Frame(self, bg=("#d0d7e2" if is_day_theme else "#16365f"), width=2)
        self._connector.grid(row=0, column=1, sticky="ns", padx=0, pady=12)

        self.right = RoundedFrame(
            self, radius=18, bg=self.bg, border=self.border,
            shadow=True, shadow_offset=(3, 4), shadow_stipple=shadow_stipple
        )
        self.right.grid(row=0, column=2, sticky="nsew", padx=(8, 0), pady=0)

        self.f_title = tkfont.Font(family="Segoe UI", size=14, weight="bold")
        self.f_status = tkfont.Font(family="Segoe UI", size=20, weight="bold")
        self.f_sub = tkfont.Font(family="Segoe UI", size=16, weight="bold")
        self.f_lines = tkfont.Font(family="Segoe UI", size=15, weight="bold")

        self.title = tk.Label(self.left, text="CLUBE", font=self.f_title,
                              fg="#ffffff", bg="#2e7d32", padx=14, pady=7, anchor="w")
        self.status = tk.Label(self.left, text="ABERTO AGORA", font=self.f_status,
                               fg="#ffffff", bg="#43a047", padx=14, pady=9, anchor="w")
        self.sub = tk.Label(self.left, text="FECHA ÀS 22H", font=self.f_sub,
                            fg="#7a1111", bg="#f3e88f", padx=14, pady=9, anchor="w")

        self.lines = tk.Label(self.right, text="", font=self.f_lines,
                              fg="#27364a" if is_day_theme else "#d8e6ff",
                              bg=self.bg, justify="left", anchor="nw")

        self._place_left_job = None
        self._place_right_job = None
        self._last_left_wh = (0, 0)
        self._last_right_wh = (0, 0)

        self.left.bind("<Configure>", self._on_left_cfg)
        self.right.bind("<Configure>", self._on_right_cfg)

        self._mode = 0
        self._items = self._build_items()
        self.update_view()

    def _build_items(self):
        return [
            {
                "name": "CLUBE",
                "rules": {
                    "SEG": ("06:00", "22:00"),
                    "TER": ("06:00", "22:00"),
                    "QUA": ("06:00", "22:00"),
                    "QUI": ("06:00", "22:00"),
                    "SEX": ("06:00", "22:00"),
                    "SAB": ("08:00", "20:00"),
                    "DOM": ("09:00", "20:00"),
                },
                "lines": "SEG–QUI  6H–22H\nSEX      6H–22H\nSÁB      8H–20H\nDOM/FER  9H–20H",
            },
            {
                "name": "SECRETARIA",
                "rules": {
                    "SEG": ("08:00", "20:00"),
                    "TER": ("08:00", "20:00"),
                    "QUA": ("08:00", "20:00"),
                    "QUI": ("08:00", "20:00"),
                    "SEX": ("08:00", "20:00"),
                    "SAB": ("08:00", "20:00"),
                    "DOM": ("09:00", "20:00"),
                },
                "lines": "SEG–QUI  8H–20H\nSEX      8H–20H\nSÁB      8H–20H\nDOM/FER  9H–20H",
            },
            {
                "name": "ACADEMIA",
                "rules": {
                    "SEG": ("06:00", "21:30"),
                    "TER": ("06:00", "21:30"),
                    "QUA": ("06:00", "21:30"),
                    "QUI": ("06:00", "21:30"),
                    "SEX": ("06:00", "21:00"),
                    "SAB": ("08:00", "12:00"),
                    "DOM": (None, None),
                },
                "lines": "SEG–QUI  06:00–21:30\nSEX      06:00–21:00\nSÁB      08:00–12:00\nDOM      FECHADO",
            }
        ]

    def _on_left_cfg(self, _evt=None):
        w = self.left.winfo_width()
        h = self.left.winfo_height()
        if (w, h) == self._last_left_wh:
            return
        self._last_left_wh = (w, h)
        _debounce_after(self.left, "_place_left_job", 33, self._place_left)

    def _on_right_cfg(self, _evt=None):
        w = self.right.winfo_width()
        h = self.right.winfo_height()
        if (w, h) == self._last_right_wh:
            return
        self._last_right_wh = (w, h)
        _debounce_after(self.right, "_place_right_job", 33, self._place_right)

    def _place_left(self, _evt=None):
        self.left.delete("win")
        w = self.left.winfo_width()
        h = self.left.winfo_height()
        if w <= 10 or h <= 10:
            return

        scale = max(0.78, min(1.0, h / 200.0))
        self.f_title.configure(size=max(11, int(14 * scale)))
        self.f_status.configure(size=max(14, int(20 * scale)))
        self.f_sub.configure(size=max(12, int(16 * scale)))

        pad = max(12, int(18 * scale))
        gap = max(4, int(6 * scale))

        title_h = max(28, int(40 * scale))
        status_h = max(40, int(58 * scale))
        sub_h = max(34, int(48 * scale))

        self.title.configure(padx=max(10, int(14 * scale)), pady=max(4, int(7 * scale)))
        self.status.configure(padx=max(10, int(14 * scale)), pady=max(5, int(9 * scale)))
        self.sub.configure(padx=max(10, int(14 * scale)), pady=max(5, int(9 * scale)))

        y = max(10, int(14 * scale))
        self.left.create_window(pad, y, anchor="nw", window=self.title,
                                width=w - 2 * pad, height=title_h, tags=("win",))
        y += title_h + gap
        self.left.create_window(pad, y, anchor="nw", window=self.status,
                                width=w - 2 * pad, height=status_h, tags=("win",))
        y += status_h + gap
        self.left.create_window(pad, y, anchor="nw", window=self.sub,
                                width=w - 2 * pad, height=sub_h, tags=("win",))

    def _place_right(self, _evt=None):
        self.right.delete("win")
        w = self.right.winfo_width()
        h = self.right.winfo_height()
        if w <= 10 or h <= 10:
            return

        scale_w = max(0.95, min(1.35, w / 210.0))
        self.f_lines.configure(size=max(14, int(16 * scale_w)))

        pad = max(10, int(12 * scale_w))
        top = max(10, int(12 * scale_w))

        self.right.create_window(
            pad, top, anchor="nw", window=self.lines,
            width=w - 2 * pad,
            height=h - top - pad,
            tags=("win",)
        )

    def tick_rotate(self):
        self._mode = (self._mode + 1) % len(self._items)
        self.update_view()

    def update_view(self):
        item = self._items[self._mode]
        self.title.configure(text=item["name"])
        self.lines.configure(text=item["lines"])
        self.update_status_only()

    def update_status_only(self):
        item = self._items[self._mode]

        day = today_3letters_noaccent()
        now = time.localtime()
        now_min = now.tm_hour * 60 + now.tm_min

        open_s, close_s = item["rules"].get(day, (None, None))
        open_m = parse_hhmm(open_s) if open_s else None
        close_m = parse_hhmm(close_s) if close_s else None

        if open_m is None or close_m is None:
            is_open = False
        else:
            is_open = (open_m <= now_min < close_m)

        if is_open:
            self.status.configure(text="ABERTO AGORA", bg="#43a047")
            self.sub.configure(
                text=f"FECHA ÀS {fmt_hhmm(close_m)}",
                fg="#7a1111",
                bg="#f3e88f"
            )
        else:
            self.status.configure(text="FECHADO AGORA", bg="#3d556d")
            if open_m is not None:
                self.sub.configure(
                    text=f"ABRE ÀS {fmt_hhmm(open_m)}",
                    fg="#ffffff",
                    bg="#2b3f55"
                )
            else:
                self.sub.configure(text="SEM ATENDIMENTO", fg="#ffffff", bg="#2b3f55")


class WeatherCard(tk.Frame):
    """
    Card de clima (UI-only).
    - Não altera backend (weather.py)
    - set_weather(city, WeatherResult)
    Anti-flicker: layout simples, sem reflow agressivo.
    """
    def __init__(self, master, is_day_theme: bool):
        super().__init__(master, bd=0, highlightthickness=0)
        self.is_day_theme = is_day_theme

        try:
            self.configure(bg=master["bg"])
        except Exception:
            pass

        self.bg = "#ffffff" if is_day_theme else "#0f1a2a"
        self.border = "#d7dee8" if is_day_theme else "#20324d"
        shadow_stipple = "gray18" if is_day_theme else "gray10"

        self.panel = RoundedFrame(
            self, radius=18, bg=self.bg, border=self.border,
            shadow=True, shadow_offset=(3, 4), shadow_stipple=shadow_stipple
        )
        self.panel.pack(fill="both", expand=True)

        self.inner = tk.Frame(self.panel, bg=self.bg)
        self._inner_job = None
        self._last_wh = (0, 0)
        self.panel.bind("<Configure>", self._on_cfg)

        self.city_lbl = tk.Label(self.inner, text="CLIMA", font=("Segoe UI", 12, "bold"),
                                 fg="#1f7ab8" if is_day_theme else "#4aa3ff",
                                 bg=self.bg, anchor="w")
        self.city_lbl.pack(anchor="w", padx=14, pady=(12, 2))

        mid = tk.Frame(self.inner, bg=self.bg)
        mid.pack(fill="x", padx=14, pady=(4, 0))

        self.icon_lbl = tk.Label(mid, bg=self.bg)
        self.icon_lbl.pack(side="left", padx=(0, 10))

        self.temp_lbl = tk.Label(mid, text="—°C", font=("Segoe UI", 26, "bold"),
                                 fg="#0b2d4d" if is_day_theme else "#eaf2ff",
                                 bg=self.bg, anchor="w")
        self.temp_lbl.pack(side="left", anchor="w")

        self.desc_lbl = tk.Label(self.inner, text="—", font=("Segoe UI", 12, "bold"),
                                 fg="#5c6f86" if is_day_theme else "#b9c7dd",
                                 bg=self.bg, anchor="w", justify="left", wraplength=360)
        self.desc_lbl.pack(anchor="w", padx=14, pady=(6, 10))

        self._icon_img = None

    def _on_cfg(self, _evt=None):
        w = self.panel.winfo_width()
        h = self.panel.winfo_height()
        if (w, h) == self._last_wh:
            return
        self._last_wh = (w, h)
        _debounce_after(self.panel, "_inner_job", 33, self._place_inner)

    def _place_inner(self):
        w = self.panel.winfo_width()
        h = self.panel.winfo_height()
        if w <= 10 or h <= 10:
            return
        pad = 6
        self.panel.delete("inner")
        self.panel.create_window(
            pad, pad, anchor="nw", window=self.inner,
            width=max(10, w - 2 * pad),
            height=max(10, h - 2 * pad),
            tags=("inner",)
        )

    def _set_icon(self, base: str):
        p = _img_path_try(base)
        if not p:
            p = _img_path_try("clouds")
        if not p:
            self.icon_lbl.configure(image=None)
            self._icon_img = None
            return
        try:
            img = tk.PhotoImage(file=p)
            # padroniza tamanho (aprox 56–72) sem ficar enorme
            if img.width() > 90:
                factor = max(1, img.width() // 72)
                img = img.subsample(factor, factor)
            self._icon_img = img
            self.icon_lbl.configure(image=self._icon_img)
        except Exception:
            self.icon_lbl.configure(image=None)
            self._icon_img = None

    def set_weather(self, city: str, res: Any):
        # Extrai campos de forma robusta sem depender do formato exato do WeatherResult
        # prioriza "agora"
        now_obj = _get_any(res, "now", "current", "agora", default=res)

        temp = _get_any(now_obj, "temp_c", "temperature", "air_temperature", "temp", default=None)
        sym = _get_any(now_obj, "symbol_code", "symbol", "icon", "weather_symbol", default=None)
        desc = _get_any(now_obj, "summary", "description", "desc", "text", default=None)

        # fallback: às vezes vem em res direto
        if temp is None:
            temp = _get_any(res, "temp_c", "temperature", "air_temperature", "temp", default=None)
        if sym is None:
            sym = _get_any(res, "symbol_code", "symbol", "icon", "weather_symbol", default=None)
        if desc is None:
            desc = _get_any(res, "summary", "description", "desc", "text", default=None)

        try:
            if temp is not None:
                # evita "21.3" virar texto feio
                tval = float(temp)
                temp_s = f"{int(round(tval))}°C"
            else:
                temp_s = "—°C"
        except Exception:
            temp_s = f"{temp}°C" if temp is not None else "—°C"

        icon_base = map_symbol_to_icon(sym)
        self._set_icon(icon_base)

        self.city_lbl.configure(text=str(city).upper())
        self.temp_lbl.configure(text=temp_s)
        self.desc_lbl.configure(text=str(desc).strip().upper() if desc else "—")


# -------------------------
# Main App
# -------------------------

class SALApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("SAL - SESI Agenda Live")
        self.attributes("-fullscreen", True)
        self.configure(bg="#f5f7fb")

        self.is_day_theme = theme_is_day()
        self._apply_theme()

        self.all_items: List[ClassItem] = []
        self.last_excel_mtime: Optional[float] = None

        self.weather_last_fetch = 0.0
        self.weather_res: Optional[Any] = None
        self._weather_lock = threading.Lock()
        self._weather_inflight = False

        self._last_zero_agenda_log_ts = 0.0
        self._last_housekeeping_ts = 0.0

        # UI: garante apenas 1 loop _tick ativo
        self._tick_job = None
        self._tick_running = False

        self.bind("<Escape>", lambda e: self.destroy())

        _rotate_logs_if_needed(logger=log)

        log(f"[BOOT] APP_DIR={APP_DIR}")
        log(f"[BOOT] DATA_DIR={DATA_DIR}")
        log(f"[BOOT] EXCEL_PATH={EXCEL_PATH} exists={os.path.exists(EXCEL_PATH)}")
        log(f"[BOOT] GRAPHICS_DIR={GRAPHICS_DIR} exists={os.path.exists(GRAPHICS_DIR)}")
        log(f"[BOOT] LOG_PATH={LOG_PATH}")
        log(f"[BOOT] {SAL_UI_BUILD}")

        try:
            weather_mod.housekeeping(app_dir=APP_DIR, logger=log)
        except Exception as e:
            log(f"[WEATHER] Housekeeping error {type(e).__name__}: {e}")

        self._build_ui()

        self._reload_excel_if_needed(force=True)
        self._tick()
        self.after(9000, self._rotate_hours)

    def _apply_theme(self):
        self.is_day_theme = theme_is_day()
        self.bg_root = "#f5f7fb" if self.is_day_theme else "#06101f"
        self.bg_header = "#f5f7fb" if self.is_day_theme else "#06101f"
        self.fg_primary = "#0b2d4d" if self.is_day_theme else "#eaf2ff"
        self.fg_soft = "#5c6f86" if self.is_day_theme else "#b9c7dd"
        self.border_soft = "#cfd8e5" if self.is_day_theme else "#132744"

        self.div_hi = "#e9eef5" if self.is_day_theme else "#16365f"
        self.div_lo = "#b7c3d4" if self.is_day_theme else "#08172c"

        self.configure(bg=self.bg_root)

    def _refresh_logo(self):
        base = "logo_day" if self.is_day_theme else "logo_night"
        p = _img_path_try(base)
        if not p:
            p = _img_path_try("logo_sesi") or _img_path_try("logo_day") or _img_path_try("logo_night")
        if p:
            try:
                img = tk.PhotoImage(file=p)
                if img.width() > 320:
                    factor = max(1, img.width() // 300)
                    img = img.subsample(factor, factor)
                self.logo_img = img
                self.logo_lbl.configure(image=self.logo_img)
            except Exception:
                self.logo_lbl.configure(image=None)

    def _calc_header_geometry(self):
        sw = max(1280, self.winfo_screenwidth())
        sh = max(720, self.winfo_screenheight())

        card_h = int(sh * 0.15)
        card_h = max(140, min(175, card_h))

        gap = 12

        right_w = int(sw * 0.36)
        right_w = max(520, min(680, right_w))

        weather_w = int((right_w - gap) * 0.56)
        hours_w = (right_w - gap) - weather_w

        center_w = sw - right_w - 380
        center_w = max(460, min(650, center_w))

        min_center = 460
        if center_w < min_center:
            shrink = (min_center - center_w)
            right_w = max(500, right_w - shrink)
            weather_w = int((right_w - gap) * 0.56)
            hours_w = (right_w - gap) - weather_w
            center_w = min_center

        return center_w, right_w, hours_w, weather_w, card_h

    def _build_ui(self):
        for w in list(self.winfo_children()):
            try:
                w.destroy()
            except Exception:
                pass

        center_w, right_w, hours_w, weather_w, card_h = self._calc_header_geometry()

        self.header = tk.Frame(self, bg=self.bg_header, height=card_h + 24)
        self.header.pack(fill="x", side="top")
        self.header.pack_propagate(False)

        self.header.grid_columnconfigure(0, weight=0)
        self.header.grid_columnconfigure(1, weight=1)
        self.header.grid_columnconfigure(2, weight=0)

        # LOGO
        self.logo_img = None
        self.logo_lbl = tk.Label(self.header, bg=self.bg_header)
        self.logo_lbl.grid(row=0, column=0, sticky="w", padx=18, pady=12)

        # CENTRO
        center = tk.Frame(self.header, bg=self.bg_header, width=center_w, height=card_h)
        center.grid(row=0, column=1, sticky="nsew", padx=8, pady=10)
        center.grid_propagate(False)

        self.date_lbl = tk.Label(center, text="", font=("Segoe UI", 18, "bold"),
                                 fg=self.fg_primary, bg=self.bg_header, anchor="w", width=10)
        self.time_lbl = tk.Label(center, text="", font=("Segoe UI", 34, "bold"),
                                 fg=self.fg_primary, bg=self.bg_header, anchor="w", width=8)
        self.wd_lbl = tk.Label(center, text="", font=("Segoe UI", 16, "bold"),
                               fg=self.fg_soft, bg=self.bg_header, anchor="w", width=20)
        self.build_lbl = tk.Label(center, text=SAL_UI_BUILD, font=("Segoe UI", 10, "bold"),
                                  fg=self.fg_soft, bg=self.bg_header, anchor="w", width=22)

        self.date_lbl.pack(anchor="w", pady=(6, 0))
        self.time_lbl.pack(anchor="w", pady=(0, 0))
        self.wd_lbl.pack(anchor="w", pady=(0, 2))
        self.build_lbl.pack(anchor="w", pady=(2, 0))

        # DIREITA
        self.header_right = tk.Frame(self.header, bg=self.bg_header, width=right_w, height=card_h)
        self.header_right.grid(row=0, column=2, sticky="e", padx=16, pady=10)
        self.header_right.grid_propagate(False)

        self.hours_card = HoursCard(self.header_right, is_day_theme=self.is_day_theme)
        self.weather_card = WeatherCard(self.header_right, is_day_theme=self.is_day_theme)

        self.hours_card.configure(width=hours_w, height=card_h)
        self.weather_card.configure(width=weather_w, height=card_h)
        self.hours_card.pack_propagate(False)
        self.weather_card.pack_propagate(False)

        self.hours_card.pack(side="left", padx=(0, 12), fill="y")
        self.weather_card.pack(side="left", fill="y")

        self._refresh_logo()

        # ---- DIVIDER horizontal ----
        self.div = tk.Frame(self, bg=self.bg_root, height=3)
        self.div.pack(fill="x")

        self.div_hi_line = tk.Frame(self.div, bg=self.div_hi, height=1)
        self.div_hi_line.pack(fill="x", side="top")
        self.div_lo_line = tk.Frame(self.div, bg=self.div_lo, height=2)
        self.div_lo_line.pack(fill="x", side="top")

        # ---- MAIN ----
        self.main = tk.Frame(self, bg=self.bg_root)
        self.main.pack(fill="both", expand=True)

        self.main.grid_columnconfigure(0, weight=1, uniform="main")
        self.main.grid_columnconfigure(1, weight=0)
        self.main.grid_columnconfigure(2, weight=1, uniform="main")
        self.main.grid_rowconfigure(0, weight=1)

        self.vdiv = tk.Frame(self.main, bg="#1b3a5f", width=3)
        self.vdiv.grid(row=0, column=1, sticky="ns", padx=6, pady=14)

        self.agora = SectionFrame(self.main, "AGORA", is_day_theme=self.is_day_theme)
        self.prox = SectionFrame(self.main, "PRÓXIMAS", is_day_theme=self.is_day_theme)

        self.agora.grid(row=0, column=0, sticky="nsew", padx=(16, 8), pady=(10, 14))
        self.prox.grid(row=0, column=2, sticky="nsew", padx=(8, 16), pady=(10, 14))

    def _rotate_hours(self):
        try:
            self.hours_card.tick_rotate()
        finally:
            self.after(9000, self._rotate_hours)

    def _reload_excel_if_needed(self, force: bool = False):
        try:
            mtime = os.path.getmtime(EXCEL_PATH)
            if force or self.last_excel_mtime is None or mtime != self.last_excel_mtime:
                items = load_classes_from_excel(EXCEL_PATH)
                self.all_items = items
                self.last_excel_mtime = mtime
                log(f"[XLSX] Excel carregado: {len(self.all_items)} itens. mtime={mtime}")
        except Exception as e:
            log(f"[XLSX] Falha ao carregar Excel: {type(e).__name__}: {e}")

    def _compute_now_next(self) -> Tuple[List[Tuple], List[Tuple]]:
        now_dt = datetime.now()
        window_end = now_dt + timedelta(minutes=120)

        now_list = []
        next_list = []

        discard_day = 0
        discard_time = 0
        discard_other = 0

        for it in self.all_items:
            interval, reason = _item_interval_debug(now_dt, it)
            if not interval:
                if reason == "invalid_day":
                    discard_day += 1
                elif reason == "invalid_time":
                    discard_time += 1
                else:
                    discard_other += 1
                continue

            start_dt, end_dt = interval

            if start_dt <= now_dt < end_dt:
                total = (end_dt - start_dt).total_seconds()
                done = (now_dt - start_dt).total_seconds()
                progress = 0.0 if total <= 0 else max(0.0, min(1.0, done / total))
                now_list.append((it.start, it.end, it.activity, it.teacher, it.location, it.tag,
                                 progress, end_dt, start_dt))

            elif now_dt <= start_dt <= window_end:
                total_win = (window_end - now_dt).total_seconds()
                remaining = (start_dt - now_dt).total_seconds()
                if total_win <= 0:
                    progress_next = 0.0
                else:
                    progress_next = 1.0 - (remaining / total_win)
                    progress_next = max(0.0, min(1.0, progress_next))

                progress_next = max(0.02, progress_next)
                next_list.append((it.start, it.end, it.activity, it.teacher, it.location, it.tag,
                                  progress_next, end_dt, start_dt))

        now_list.sort(key=lambda t: t[7])   # end_dt
        next_list.sort(key=lambda t: t[8])  # start_dt

        now_cards = [(a, b, c, d, e, f, g) for (a, b, c, d, e, f, g, _end, _start) in now_list]
        next_cards = [(a, b, c, d, e, f, g) for (a, b, c, d, e, f, g, _end, _start) in next_list]

        if self.all_items and (len(now_cards) == 0 and len(next_cards) == 0):
            if time.time() - self._last_zero_agenda_log_ts > 60:
                self._last_zero_agenda_log_ts = time.time()
                sample = self.all_items[:4]
                log(
                    "[AGENDA] 0 em AGORA/PRÓXIMAS | "
                    f"now={now_dt.strftime('%Y-%m-%d %H:%M:%S')} window_end={window_end.strftime('%Y-%m-%d %H:%M:%S')} "
                    f"today_code={today_3letters_noaccent()} "
                    f"discard_day={discard_day} discard_time={discard_time} discard_other={discard_other} "
                    f"sample_items={sample}"
                )

        return now_cards[:6], next_cards[:6]

    def _weather_worker(self):
        try:
            res = weather_mod.get_weather(
                city_label="Alfenas",
                lat=-21.4267,
                lon=-45.9470,
                app_dir=APP_DIR,
                user_agent="SAL-SESIAgendaLive/2.0 (contact: local)",
                logger=log,
            )
            with self._weather_lock:
                self.weather_res = res
                self.weather_last_fetch = time.time()
        except Exception as e:
            log(f"[WEATHER] Worker error {type(e).__name__}: {e}")
        finally:
            with self._weather_lock:
                self._weather_inflight = False

    def _tick_weather(self):
        if (time.time() - self.weather_last_fetch) < 600 and self.weather_res is not None:
            return

        with self._weather_lock:
            if self._weather_inflight:
                return
            self._weather_inflight = True

        th = threading.Thread(target=self._weather_worker, daemon=True)
        th.start()

    def _tick_housekeeping(self):
        if time.time() - self._last_housekeeping_ts < 86400:
            return
        self._last_housekeeping_ts = time.time()
        try:
            _rotate_logs_if_needed(logger=log)
            weather_mod.housekeeping(app_dir=APP_DIR, logger=log)
        except Exception as e:
            log(f"[HK] error {type(e).__name__}: {e}")

    def _tick(self):
        if getattr(self, "_tick_running", False):
            return
        self._tick_running = True

        try:
            new_theme = theme_is_day()
            if new_theme != self.is_day_theme:
                self._apply_theme()
                self._build_ui()

            d, t, wd = date_time_strings()
            self.date_lbl.configure(text=d)
            self.time_lbl.configure(text=t)
            self.wd_lbl.configure(text=wd)

            self.hours_card.update_status_only()

            self._reload_excel_if_needed(force=False)

            now_cards, next_cards = self._compute_now_next()
            self.agora.set_cards(now_cards)
            self.prox.set_cards(next_cards)

            self._tick_weather()

            with self._weather_lock:
                res = self.weather_res
            if res:
                self.weather_card.set_weather("Alfenas", res)

            self._tick_housekeeping()

        except Exception:
            log("Tick error:\n" + traceback.format_exc())

        finally:
            self._tick_running = False

            try:
                if getattr(self, "_tick_job", None) is not None:
                    self.after_cancel(self._tick_job)
            except Exception:
                pass

            self._tick_job = self.after(1000, self._tick)


if __name__ == "__main__":
    try:
        SALApp().mainloop()
    except Exception:
        log("Fatal error:\n" + traceback.format_exc())
